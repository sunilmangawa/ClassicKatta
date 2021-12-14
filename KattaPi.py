#predifined liabraries
import os
import sys
sys.path.append('/home/pi/.local/lib/python3.10/site-packages/gpiozero/__init__.py')
from gpiozero import LED, Button
from time import sleep
import time
from datetime import date, datetime
import openpyxl
from openpyxl import Workbook
import RPi.GPIO as GPIO
from threading import Thread
import signal
#from katta_Backup import katta_counter
#from Time_auto import timeauto

import mysql.connector

mydb = mysql.connector.connect(
  host="localhost",
  user="coder",
  password="@Work/Minute",
  database="classicmill1"
)

curA = mydb.cursor(buffered=True)
# curA.execute("SHOW TABLES")
# for x in curA:
#   print(x)

# sql = "INSERT INTO cm1kattatb (katta, startfill, endfill, filltime, date, avgkatta, machineon, machineoff) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
sql = "INSERT INTO cm1kattatable(katta, startfill, endfill, filltime, date, avgkatta, machineon, machineoff) VALUES(%s, %s, %s, %s, %s, %s, %s, %s)"



# Excel file setting to save katta & time
wb=Workbook()
ws = wb.active
ws['A1'] = 'Katta_Number'
ws['B1'] = 'Fill_Time'
ws['C1'] = 'Date_Time'



# PIN Setting for Katta Button & LED
btn_katta_fill = Button(5)
led = LED(22) #GPIO22  PIN NO 22
led.source=btn_katta_fill


# Auto timer PIN setting for Vibrator & Conveyar
GPIO.setwarnings(False)

GPIO.setmode(GPIO.BCM)

GPIO.setup(13,GPIO.OUT)
GPIO.output(13,GPIO.HIGH)

GPIO.setup(19,GPIO.OUT)
GPIO.output(19,GPIO.HIGH)

GPIO.setup(26,GPIO.OUT)
GPIO.output(26,GPIO.HIGH)

GPIO.setup(21,GPIO.OUT)
GPIO.output(21,GPIO.HIGH)
count=100


counted_bag = 0

machine_on = datetime.now()
machine_start = machine_on.strftime(f"%Y-%m-%d %H:%M:%S")
# katta_Date = machine_on.strftime("%d %B %Y")
# Katta Counter code to Count filled bags & their time
def katta_counter():
    
    global counted_bag
    b,j=2,3
    time_taken=0
    try:
        global machine_start, curA, sql
        startTime=time.time()
        dateToday=date.today()

        while True:
            if btn_katta_fill.is_pressed:
                btn_katta_fill.wait_for_release()
                t1 = time.time()
                kattaTimeOn = datetime.now().strftime("%H:%M:%S")
                sleep(6)
                #if (round(time.time()-t1),2)>7:
                machine_off = datetime.now().strftime(f"%Y-%m-%d %H:%M:%S")
                btn_katta_fill.wait_for_press()
                counted_bag += 1
                t2 = time.time()
                kattaTimeOff = datetime.now().strftime("%H:%M:%S")
                time_taken = round(t2-t1,2)
                averageKatta=round(((counted_bag/(t2-startTime))*3600),2)
                print("Katta no.: ", counted_bag ," in :", time_taken ,"seconds -|- Avg/hr = ", averageKatta)
                sleep(.1)
                # sql = "INSERT INTO cm1KattaTable (katta, startfill, endfill, filltime, date, avgkatta, machineon, machineoff) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
                val = (counted_bag, kattaTimeOn, kattaTimeOff, time_taken, dateToday, averageKatta, machine_start, machine_off)
                curA.execute(sql, val)
                mydb.commit()
                print(curA.rowcount, "record inserted.")
                sleep(.1)
                
            else:
                btn_katta_fill.wait_for_press()
            
            for i in range(b,j):
                #if ws[f'A{i}']==True and (ws[f'B{i}']==True and ws[f'C{i}']==True)
                ws[f'A{i}']=counted_bag
                ws[f'B{i}']=time_taken
                ws[f'C{i}']=datetime.now()
                #else:
                 #   b+=1
                  #  j+=1
                   # continue
            wb.save('Katta_Counter.xlsx')
            #print('Excel file created Successfully')
            b+=1
            j+=1
    except KeyboardInterrupt:
        print("Programmed stopped by pressing: CTRL+C")

# Auto timer function to control Vibrator & Conveyar
def timeauto():
    try:
        sleep(360)
        while True:
            count <=200
            print("System start")
            GPIO.output(21,GPIO.LOW)
            print("vibrator on")
            time.sleep(2)
            GPIO.output(21,GPIO.HIGH)
            time.sleep(2)
            GPIO.output(19,GPIO.LOW)   #mall pin on
            print("mall on")
            time.sleep(2)
            GPIO.output(19,GPIO.HIGH) # mallpin low
            time.sleep(16)
            GPIO.output(13,GPIO.LOW)
            print("mall off")
            time.sleep(2)
            GPIO.output(13,GPIO.HIGH)
            time.sleep(45)
        
            GPIO.output(26,GPIO.LOW)
            print("vibrator off")
            time.sleep(2)
            GPIO.output(26,GPIO.HIGH)
            time.sleep(1)
            time.sleep(80)
            GPIO.output(21,GPIO.LOW)
            print("vibrator on")
            time.sleep(2)
            GPIO.output(21,GPIO.HIGH)
            time.sleep(2)
            print("All off")
            #GPIO.cleanup()
            

    except KeyboardInterrupt:
        print("Quit")
        #GPIO.cleanup()
def exit_test():
    if sys.argv== signal.SIGINT:
        os.kill()
        raise KeyboardInterrupt

if __name__ == '__main__':
    Thread(target = katta_counter).start()
    # Thread(target = timeauto).start()
    Thread(target = exit_test).start()